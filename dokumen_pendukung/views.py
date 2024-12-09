from django.shortcuts import render

import os
import tempfile
from django.http import FileResponse, HttpResponse, HttpResponseNotFound
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import json
from rest_framework.decorators import api_view
from openpyxl.styles import Font
from .models import InvoiceDP, InvoiceFinal, KwitansiDP, KwitansiFinal
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from pptx import Presentation
from django.core.files.storage import default_storage


User = get_user_model()

@csrf_exempt
def upload_template_kontrak(request):
    if request.method == 'POST':
        if 'template' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        file = request.FILES['template']
        
        # Check if the file is a .ppt file
        if not file.name.endswith('.docx'):
            return JsonResponse({'error': 'Only .docx files are allowed'}, status=400)

        # Define the path where the file should be saved
        file_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung', 'templates', 'templateKontrak.docx')
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Delete the existing file if it exists
        if os.path.exists(file_path):
            os.remove(file_path)

        # Save the new file
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return JsonResponse({'message': 'File uploaded successfully', 'file_name': 'templateKontrak.docx'}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@api_view(['GET']) 
def download_template_kontrak(request):
    docx_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKontrak.docx')
    
    # Check if the file exists before opening it
    if os.path.exists(docx_path):
        response = FileResponse(open(docx_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename="templateKontrak.docx"'
        return response
    else:
        return HttpResponseNotFound("Template file not found.")

@csrf_exempt
def upload_template_proposal(request):
    if request.method == 'POST':
        if 'template' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        file = request.FILES['template']
        
        # Check if the file is a .ppt file
        if not file.name.endswith('.ppt'):
            return JsonResponse({'error': 'Only .ppt files are allowed'}, status=400)

        # Define the path where the file should be saved
        file_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung', 'templates', 'templateProposal.ppt')
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Delete the existing file if it exists
        if os.path.exists(file_path):
            os.remove(file_path)

        # Save the new file
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        return JsonResponse({'message': 'File uploaded successfully', 'file_name': 'templateProposal.ppt'}, status=200)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@api_view(['GET']) 
def download_template_proposal(request):
    if request.method == 'GET':
        ppt_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateProposal.ppt')
        
        if os.path.exists(ppt_path):
            response = FileResponse(open(ppt_path, 'rb'), content_type='application/vnd.ms-powerpoint')
            response['Content-Disposition'] = 'attachment; filename="templateProposal.ppt"'
            return response
        else:
            return HttpResponseNotFound("Template file not found.")
    
    # If not GET, return 405 directly
    return JsonResponse({"detail": "Method not allowed"}, status=405)

# Helper function to convert month to Roman numeral
def month_to_roman(month):
    roman_numerals = {
        1: "I", 2: "II", 3: "III", 4: "IV",
        5: "V", 6: "VI", 7: "VII", 8: "VIII",
        9: "IX", 10: "X", 11: "XI", 12: "XII"
    }
    return roman_numerals.get(month, "")

# Helper function to get the next invoice number
def get_next_invoice_number():
    current_date = datetime.now()
    year = current_date.year
    counter_file_path = os.path.join(settings.BASE_DIR, 'invoice_{year}_counter.txt')

    # Check if the file exists, and create it if not
    if not os.path.exists(counter_file_path):
        with open(counter_file_path, 'w') as file:
            file.write("0")  # Initialize counter at 0

    # Read the current counter value
    with open(counter_file_path, 'r') as file:
        last_number = int(file.read())

    # Increment and format the number with leading zeros
    new_number = last_number + 1
    formatted_number = f"{new_number:03}"

    # Save the updated number back to the file
    with open(counter_file_path, 'w') as file:
        file.write(str(new_number))

    return formatted_number


# Helper function to get the next kwitansi number
def get_next_kwitansi_number():
    current_date = datetime.now()
    year = current_date.year
    counter_file_path = os.path.join(settings.BASE_DIR, 'kwitansi_{year}_counter.txt')

    # Check if the file exists, and create it if not
    if not os.path.exists(counter_file_path):
        with open(counter_file_path, 'w') as file:
            file.write("0")  # Initialize counter at 0

    # Read the current counter value
    with open(counter_file_path, 'r') as file:
        last_number = int(file.read())

    # Increment and format the number with leading zeros
    new_number = last_number + 1
    formatted_number = f"{new_number:03}"

    # Save the updated number back to the file
    with open(counter_file_path, 'w') as file:
        file.write(str(new_number))

    return formatted_number



@api_view(['POST'])
def generate_invoice_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Generate unique invoice code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    invoice_number = get_next_invoice_number()
    invoice_code = f"Inv No: {invoice_number}/SURV/LSI/{month_roman}/{year}"
    invoice_id = f"{invoice_number}/SURV/LSI/{month_roman}/{year}"
    doc = "invoiceDP"

    # Save the data to the invoice_dp table
    if not InvoiceDP.objects.filter(id=invoice_id).exists(): 
        InvoiceDP.objects.create(
            id=invoice_id,
            client_name=user_data['client_name'],
            survey_name=user_data['survey_name'],
            respondent_count=user_data['respondent_count'],
            address=user_data['address'],
            amount=user_data['amount'],
            nominal_tertulis=user_data['nominal_tertulis'],
            paid_percentage=user_data['paid_percentage'],
            additional_info=user_data['additional_info'],
            date=user_data['date'],
            doc_type=doc
        )
    else:
        print(f"Record with id {invoice_id} already exists. Skipping creation.")

    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceDP.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = invoice_code
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    sheet.add_image(ttd_img, 'G37')

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceDP_{invoice_code}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def export_existing_invoice_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'invoice_code': data.get('id', ''),
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceDP.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = user_data['invoice_code']
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    sheet.add_image(ttd_img, 'G37')

    invoice_id = f"Inv No: {user_data['invoice_code']}"

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceDP_{invoice_id}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def export_existing_invoice_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'invoice_code': data.get('id', ''),
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceFinal.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = user_data['invoice_code']
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    sheet.add_image(ttd_img, 'G37')

    invoice_id = f"Inv No: {user_data['invoice_code']}"

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceFinal_{invoice_id}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def generate_invoice_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
   
    user_data = {
        'client_name': data.get('client_name', 'Default Client'),
        'survey_name': data.get('survey_name', 'Default Survey'),
        'respondent_count': data.get('respondent_count', '0'),
        'address': data.get('address', 'Default Address'),
        'amount': data.get('amount', '0'),
        'paid_percentage': data.get('paid_percentage', '60'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', 'No additional info'),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    # Generate unique invoice code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    invoice_number = get_next_invoice_number()
    invoice_code = f"Inv No: {invoice_number}/SURV/LSI/{month_roman}/{year}"
    invoice_id = f"{invoice_number}/SURV/LSI/{month_roman}/{year}"
    doc = "invoiceFinal"

    # Save the data to the invoice_dp table
    InvoiceFinal.objects.create(
        id=invoice_id,
        client_name=user_data['client_name'],
        survey_name=user_data['survey_name'],
        respondent_count=user_data['respondent_count'],
        address=user_data['address'],
        amount=user_data['amount'],
        nominal_tertulis=user_data['nominal_tertulis'],
        paid_percentage=user_data['paid_percentage'],
        additional_info=user_data['additional_info'],
        date=user_data['date'],
        doc_type=doc
    )

    # Load the Excel template
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateInvoiceFinal.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active
    respondent_message = f"sampel {user_data['respondent_count']} responden"
    paid_percentage_message = f"{user_data['paid_percentage']}%"
    formatted_date = datetime.strptime(user_data['date'], "%Y-%m-%d").strftime("%d %B %Y")
    date_message = f"Date: {formatted_date}"
    sheet.merge_cells('B31:E31')
    sheet.merge_cells('C16:F16')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('B28:E28')
    sheet.merge_cells('G14:H14')

    # Fill out the required cells in the Excel file with user data
    sheet['A9'] = invoice_code
    sheet['C16'] = user_data['client_name']           
    sheet['B27'] = user_data['survey_name']           
    sheet['B31'] = respondent_message      
    sheet['C18'] = user_data['address']               
    sheet['G27'] = user_data['amount']                
    sheet['C35'] = user_data['nominal_tertulis']
    sheet['C35'].font = Font(name="Times New Roman", bold=True, underline="single")
    sheet['F27'] = paid_percentage_message      
    sheet['B28'] = user_data['additional_info']       
    sheet['G14'] = date_message                 

    # Path to the image you want to add
    header_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/header.png')  
    header_img = Image(header_image_path)
    header_img.width, header_img.height = 846.6, 136  

    # Path to the image you want to add
    invoice_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/invoice.png') 
    invoice_img = Image(invoice_image_path)
    invoice_img.width, invoice_img.height = 275.9, 75.59

    # Path to the image you want to add
    ttd_image_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/images/ttd.png')
    ttd_img = Image(ttd_image_path)
    ttd_img.width, ttd_img.height = 314.83, 173.48

    # Add image to the specified cell location
    sheet.add_image(header_img, 'A1') 
    sheet.add_image(invoice_img, 'G8')
    sheet.add_image(ttd_img, 'G37')

    # Generate a filename
    filename = f"{user_data['survey_name']}_invoiceFinal_{invoice_code}.xlsx"

    # Prepare the response as an Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to the response
    workbook.save(response)
    return response

    return HttpResponse(status=405)  # Method not allowed for non-POST requests

@api_view(['POST'])
def generate_kwitansi_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'pembayar': data.get('pembayar', ''),
        'tujuan_pembayaran': data.get('tujuan_pembayaran', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }

    # Generate unique kwitansi code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    kwitansi_number = get_next_kwitansi_number()
    kwitansi_code = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    kwitansi_id = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    doc = "kwitansiDP"

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    # Save the data to the kwitansi_dp table
    KwitansiDP.objects.create(
        id=kwitansi_id,
        # pembayar=user_data['pembayar'],
        # tujuan_pembayaran=user_data['tujuan_pembayaran'],
        client_name=user_data['pembayar'],
        survey_name=user_data['tujuan_pembayaran'],
        nominal_tertulis=user_data['nominal_tertulis'],
        additional_info=user_data['additional_info'],
        amount=user_data['amount'],
        date=user_data['date'],
        doc_type=doc
    )

    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = kwitansi_code
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiDP_{kwitansi_code}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response

@api_view(['POST'])
def export_existing_kwitansi_dp(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'kwitansi_code': data.get('id', ''),
        'pembayar': data.get('client_name', ''),
        'tujuan_pembayaran': data.get('survey_name', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")


    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = user_data['kwitansi_code']
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiDP_{user_data['kwitansi_code']}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response

@api_view(['POST'])
def export_existing_kwitansi_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'kwitansi_code': data.get('id', ''),
        'pembayar': data.get('client_name', ''),
        'tujuan_pembayaran': data.get('survey_name', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }


    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")


    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = user_data['kwitansi_code']
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiFinal_{user_data['kwitansi_code']}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response


@api_view(['POST'])
def generate_kwitansi_final(request):
    if request.method == 'POST':
        # Parse JSON data from the request body
        data = json.loads(request.body)
    
    user_data = {
        'pembayar': data.get('pembayar', ''),
        'tujuan_pembayaran': data.get('tujuan_pembayaran', ''),
        'amount': data.get('amount', '0'),
        'nominal_tertulis' : data.get('nominal_tertulis', ''),
        'additional_info': data.get('additional_info', ''),
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
    }

    # Generate unique kwitansi code
    current_date = datetime.now()
    month_roman = month_to_roman(current_date.month)
    year = current_date.year
    kwitansi_number = get_next_kwitansi_number()
    kwitansi_code = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    kwitansi_id = f"{kwitansi_number}/IDR-KWT/{month_roman}/{year}"
    doc = "kwitansiFinal"

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    formatted_date = datetime.strptime(user_data['date'], '%Y-%m-%d').strftime("Jakarta, %d %B %Y")

    # Save the data to the kwitansi_dp table
    KwitansiDP.objects.create(
        id=kwitansi_id,
        # pembayar=user_data['pembayar'],
        # tujuan_pembayaran=user_data['tujuan_pembayaran'],
        client_name=user_data['pembayar'],
        survey_name=user_data['tujuan_pembayaran'],
        nominal_tertulis=user_data['nominal_tertulis'],
        additional_info=user_data['additional_info'],
        amount=user_data['amount'],
        date=user_data['date'],
        doc_type=doc
    )

    # load excel template for kwitansi dp
    template_path = os.path.join(settings.BASE_DIR, 'dokumen_pendukung/templates/templateKwitansi.xlsx')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    times_new_roman_font = Font(name="Times New Roman")
    bold_times_new_roman_font = Font(name="Times New Roman", bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    for cell in sheet['A']:
        cell.font = bold_times_new_roman_font

    sheet.merge_cells('A11:L11')
    sheet.merge_cells('E14:G14')
    sheet.merge_cells('E16:G16')
    sheet.merge_cells(start_row=17, start_column=1, end_row=18, end_column=3)
    sheet.merge_cells(start_row=17, start_column=5, end_row=18, end_column=7)
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('B27:E27')
    sheet.merge_cells('K27:L27')

    # fill cells with input from user_data
    sheet['A11'] = kwitansi_code
    sheet['E14'] = user_data['pembayar']
    sheet['E16'] = f"# {user_data['nominal_tertulis']} #"
    sheet['E17'] = user_data['tujuan_pembayaran']
    sheet['E19'] = user_data['additional_info']
    sheet['B27'] = user_data['amount']
    sheet['K27'] = formatted_date

    # generate file name
    filename = f"KwitansiFinal_{kwitansi_code}.xlsx"
    # response as excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # save  workbook to the response
    workbook.save(response)
    return response